"""
=========================================================
                REPORTS & ANALYTICS ROUTES
        Modernized Healthcare PBM Portal
=========================================================
"""

import csv

from openpyxl import Workbook

from openpyxl.styles import (

    Font,

    PatternFill,

    Alignment,

    Border,

    Side

)

from openpyxl.utils import get_column_letter

from io import BytesIO

from flask import send_file

from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer
)

from reportlab.lib import colors

from reportlab.lib.styles import getSampleStyleSheet

from reportlab.lib.units import inch

from datetime import datetime, timedelta

from flask import (
    Blueprint,
    render_template,
    request,
    redirect,
    url_for,
    flash,
    send_file,
    Response
)

from sqlalchemy import func

from app.extensions import db

from app.models.claim import Claim

from app.models.patient import Patient

from app.models.pharmacy import Pharmacy

from app.services.notification_service import NotificationService

from app.services.system_service import SystemService

reports_bp = Blueprint(
    "reports",
    __name__,
    url_prefix="/reports"
)


# ==========================================================
# REPORTS DASHBOARD
# ==========================================================

@reports_bp.route("/")
def index():

    # =====================================================
    # DATE FILTER
    # =====================================================

    filter_type = request.args.get(
        "filter",
        "all"
    )

    query = Claim.query

    today = datetime.utcnow()

    if filter_type == "today":

        query = query.filter(
            func.date(Claim.created_at) == today.date()
        )

    elif filter_type == "week":

        query = query.filter(
            Claim.created_at >= today - timedelta(days=7)
        )

    elif filter_type == "month":

        query = query.filter(
            Claim.created_at >= today - timedelta(days=30)
        )

    claims = query.order_by(
        Claim.created_at.desc()
    ).all()

    # =====================================================
    # STATISTICS
    # =====================================================

    total_claims = len(claims)

    approved_claims = sum(
        1
        for c in claims
        if c.status == "Approved"
    )

    pending_claims = sum(
        1
        for c in claims
        if c.status == "Pending"
    )

    rejected_claims = sum(
        1
        for c in claims
        if c.status == "Rejected"
    )

    total_amount = round(
        sum(
            c.claim_amount
            for c in claims
        ),
        2
    )

    average_amount = round(

        total_amount / total_claims,

        2

    ) if total_claims else 0

    # =====================================================
    # DASHBOARD STATISTICS
    # =====================================================

    total_patients = Patient.query.count()

    total_medicines = Pharmacy.query.count()

    total_predictions = Claim.query.filter(
        Claim.prediction_date.isnot(None)
    ).count()

    high_risk = Claim.query.filter(
        Claim.risk_level == "High"
    ).count()

    duplicate_claims = Claim.query.filter(
        Claim.duplicate_claim == True
    ).count()

    low_stock = Pharmacy.query.filter(
        Pharmacy.stock_quantity <= 10
    ).count()

    # =====================================================
    # PATIENT GENDER STATISTICS
    # =====================================================

    male_patients = Patient.query.filter(
        Patient.gender == "Male"
    ).count()

    female_patients = Patient.query.filter(
        Patient.gender == "Female"
    ).count()

    # =====================================================
    # RENDER PAGE
    # =====================================================

    return render_template(

        "reports/index.html",

        claims=claims,

        filter_type=filter_type,

        total_patients=total_patients,

        total_claims=total_claims,

        total_medicines=total_medicines,

        total_predictions=total_predictions,

        approved_claims=approved_claims,

        pending_claims=pending_claims,

        rejected_claims=rejected_claims,

        total_amount=total_amount,

        average_amount=average_amount,

        male_patients=male_patients,

        female_patients=female_patients,

        high_risk=high_risk,

        duplicate_claims=duplicate_claims,

        low_stock=low_stock

    )


    enabled = SystemService.get(

        "email_notifications",

        "Enabled"

    )

    if enabled == "Enabled":


        NotificationService.create_notification(

        title="New Report Generated",

        message="Healthcare report generated successfully.",

        category="Report",

        priority="Low",

        action_url="/reports",

        icon="fa-file-pdf",

        color="success"

    )

    # ==========================================================
    # EXPORT PDF
    # ==========================================================

@reports_bp.route("/export/pdf")
def export_pdf():

    claims = Claim.query.order_by(
        Claim.created_at.desc()
    ).all()

    buffer = BytesIO()

    document = SimpleDocTemplate(

        buffer,

        pagesize=(8.27 * inch, 11.69 * inch)

    )

    styles = getSampleStyleSheet()

    elements = []

    title = Paragraph(

        "<b>Healthcare Claims Analytics Report</b>",

        styles["Title"]

    )

    elements.append(title)

    elements.append(

        Spacer(

            1,

            0.30 * inch

        )

    )

    generated = Paragraph(

        f"Generated : {datetime.now().strftime('%d-%m-%Y %I:%M %p')}",

        styles["Normal"]

    )

    elements.append(generated)

    elements.append(

        Spacer(

            1,

            0.30 * inch

        )

    )

    total_claims = len(claims)

    approved = len(

        [

            c

            for c in claims

            if c.status == "Approved"

        ]

    )

    pending = len(

        [

            c

            for c in claims

            if c.status == "Pending"

        ]

    )

    rejected = len(

        [

            c

            for c in claims

            if c.status == "Rejected"

        ]

    )

    summary = [

        ["Total Claims", total_claims],

        ["Approved", approved],

        ["Pending", pending],

        ["Rejected", rejected]

    ]

    summary_table = Table(

        summary,

        colWidths=[220,150]

    )

    summary_table.setStyle(

        TableStyle([

            ("BACKGROUND",(0,0),(-1,0),colors.white),

            ("GRID",(0,0),(-1,-1),1,colors.grey),

            ("BACKGROUND",(0,0),(0,-1),colors.HexColor("#2563eb")),

            ("TEXTCOLOR",(0,0),(0,-1),colors.white),

            ("FONTNAME",(0,0),(-1,-1),"Helvetica-Bold"),

            ("BOTTOMPADDING",(0,0),(-1,-1),10)

        ])

    )

    elements.append(summary_table)

    elements.append(

        Spacer(

            1,

            0.35 * inch

        )

    )

    table_data = [

        [

            "Claim ID",

            "Patient",

            "Hospital",

            "Status",

            "Amount"

        ]

    ]

    for claim in claims:

        table_data.append(

            [

                claim.claim_id,

                claim.patient_id,

                claim.hospital_name,

                claim.status,

                f"₹ {claim.claim_amount:,.2f}"

            ]

        )

    report_table = Table(

        table_data,

        repeatRows=1

    )

    report_table.setStyle(

        TableStyle([

            ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#2563eb")),

            ("TEXTCOLOR",(0,0),(-1,0),colors.white),

            ("GRID",(0,0),(-1,-1),0.5,colors.grey),

            ("BACKGROUND",(0,1),(-1,-1),colors.whitesmoke),

            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),

            ("BOTTOMPADDING",(0,0),(-1,0),10)

        ])

    )

    elements.append(report_table)

    document.build(

        elements

    )

    buffer.seek(0)

    return send_file(

        buffer,

        as_attachment=True,

        download_name="Healthcare_Report.pdf",

        mimetype="application/pdf"

    )


    # ==========================================================
    # EXPORT EXCEL
    # ==========================================================

@reports_bp.route("/export/excel")
def export_excel():

    claims = Claim.query.order_by(

        Claim.created_at.desc()

    ).all()

    workbook = Workbook()

    worksheet = workbook.active

    worksheet.title = "Healthcare Report"

    # =====================================================
    # TITLE
    # =====================================================

    worksheet.merge_cells("A1:H1")

    title = worksheet["A1"]

    title.value = "Healthcare Claims Analytics Report"

    title.font = Font(

        bold=True,

        size=18

    )

    title.alignment = Alignment(

        horizontal="center"

    )

    # =====================================================
    # GENERATED DATE
    # =====================================================

    worksheet["A3"] = "Generated"

    worksheet["B3"] = datetime.now().strftime(

        "%d-%m-%Y %I:%M %p"

    )

    # =====================================================
    # HEADERS
    # =====================================================

    headers = [

        "Claim ID",

        "Patient ID",

        "Hospital",

        "Disease",

        "Status",

        "Risk Level",

        "Claim Amount",

        "Prediction Date"

    ]

    header_fill = PatternFill(

        fill_type="solid",

        fgColor="2563EB"

    )

    header_font = Font(

        color="FFFFFF",

        bold=True

    )

    border = Border(

        left=Side(style="thin"),

        right=Side(style="thin"),

        top=Side(style="thin"),

        bottom=Side(style="thin")

    )

    row = 5

    for column, header in enumerate(

        headers,

        start=1

    ):

        cell = worksheet.cell(

            row=row,

            column=column

        )

        cell.value = header

        cell.fill = header_fill

        cell.font = header_font

        cell.alignment = Alignment(

            horizontal="center"

        )

        cell.border = border

    # =====================================================
    # DATA
    # =====================================================

    row = 6

    for claim in claims:

        worksheet.cell(

            row=row,

            column=1

        ).value = claim.claim_id

        worksheet.cell(

            row=row,

            column=2

        ).value = claim.patient_id

        worksheet.cell(

            row=row,

            column=3

        ).value = claim.hospital_name

        worksheet.cell(

            row=row,

            column=4

        ).value = claim.disease

        worksheet.cell(

            row=row,

            column=5

        ).value = claim.status

        worksheet.cell(

            row=row,

            column=6

        ).value = claim.risk_level

        worksheet.cell(

            row=row,

            column=7

        ).value = claim.claim_amount

        worksheet.cell(

            row=row,

            column=8

        ).value = (

            claim.prediction_date.strftime(

                "%d-%m-%Y %H:%M"

            )

            if claim.prediction_date

            else ""

        )

        for column in range(1, 9):

            worksheet.cell(

                row=row,

                column=column

            ).border = border

        row += 1

    # =====================================================
    # AUTO WIDTH
    # =====================================================

    for column_cells in worksheet.columns:

        length = max(

            len(str(cell.value))

            if cell.value is not None

            else 0

            for cell in column_cells

        )

        worksheet.column_dimensions[

            get_column_letter(

                column_cells[0].column

            )

        ].width = min(

            length + 4,

            40

        )

    # =====================================================
    # SAVE
    # =====================================================

    buffer = BytesIO()

    workbook.save(buffer)

    buffer.seek(0)

    return send_file(

        buffer,

        as_attachment=True,

        download_name="Healthcare_Report.xlsx",

        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    )


    # ==========================================================
    # EXPORT CSV
    # ==========================================================

@reports_bp.route("/export/csv")
def export_csv():

    claims = Claim.query.order_by(
        Claim.created_at.desc()
    ).all()

    def generate():

        yield ",".join([
            "Claim ID",
            "Patient ID",
            "Hospital",
            "Disease",
            "Status",
            "Risk Level",
            "Risk Score",
            "Claim Amount",
            "Duplicate Claim",
            "Prediction Date"
        ]) + "\n"

        for claim in claims:

            prediction_date = ""

            if claim.prediction_date:

                prediction_date = claim.prediction_date.strftime(
                    "%d-%m-%Y %H:%M"
                )

            yield ",".join([

                str(claim.claim_id),

                str(claim.patient_id),

                str(claim.hospital_name),

                str(claim.disease),

                str(claim.status),

                str(claim.risk_level),

                str(claim.predicted_risk),

                str(claim.claim_amount),

                "Yes" if claim.duplicate_claim else "No",

                prediction_date

            ]) + "\n"

    return Response(

        generate(),

        mimetype="text/csv",

        headers={

            "Content-Disposition":

            "attachment; filename=Healthcare_Report.csv"

        }

    )